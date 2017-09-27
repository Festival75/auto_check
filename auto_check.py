import datetime
import json
from openpyxl import load_workbook, drawing
import openpyxl
from openpyxl.drawing.image import Image
import logging

logging.basicConfig(filename='auto_check.log', format='%(asctime)s - %(levelname)s : %(message)s', level=logging.INFO)


def read_config(path_to_config: str) -> dict:
    """Считать данные из config.json"""
    try:
        with open(path_to_config) as data_file:
            data = json.load(data_file)
        print('[+] Reading config file')
        logging.info('[+] Конфигурационный фаил прочитан')
        return data
    except Exception as err:
        logging.error('[-] Произошла ошибка в модуле read_config: {0}'.format(str(err)))
        exit(0)


def write_win_size(config: dict):
    """Записать данные о свободном месте на серверах Windows из WinDiskSpace.json в excel таблицу"""
    try:
        wdsp = config['PATH']['DISK_SPACE_FILE'] + 'WinDiskSpace.json'

        with open(wdsp) as wds_file:
            wds = json.load(wds_file)
        for ost in config['OST']:
            xlsp = config['PATH']['XLS_FILE']
            xlst = config['TYPE']['EXCEL']
            wb = load_workbook(xlsp + ost + xlst)
            w = wb.active
            w['G8'] = str(wds[ost]['ABBYY_PS_C']) + 'Gb free'
            w['G7'] = str(wds[ost]['ABBYY_DB_C']) + 'Gb free'
            w['G6'] = str(wds[ost]['ABBYY_DB_E']) + 'Gb free'
            w['G5'] = str(wds[ost]['ABBYY_AS_C']) + 'Gb free'
            w['G4'] = str(wds[ost]['ABBYY_AS_E']) + 'Gb free'
            w['G18'] = str(wds[ost]['ICC01_C']) + 'Gb free'
            w['G19'] = str(wds[ost]['ICC02_C']) + 'Gb free'
            w['C26'] = datetime.datetime.now().today()
            wb.save(xlsp + ost + xlst)
            wb.close()
            print('[+] WinSize {0} written into {1}:'.format(ost, xlsp + ost + xlst))
            logging.info('[+] Свободное место Windows серверов {0} записано в {1}'.format(ost, xlsp + ost + xlst))
    except Exception as err:
        logging.error('[-] Произошла ошибка в модуле write_win_size: {0}'.format(str(err)))
        exit(0)


def write_image(config: dict):
    """Записать PNG фаил из папки со скриншотами в excel таблицу"""
    for ost in config['OST']:
        try:
            xlsp = config['PATH']['XLS_FILE']
            xlst = config['TYPE']['EXCEL']
            imgp = config['PATH']['VWTOOL_SCREENSHOT']
            imgt = config['TYPE']['INPUT_IMAGE']
            wb = load_workbook(xlsp + ost + xlst)
            ws = wb.active
            img = openpyxl.drawing.image.Image(imgp + ost + imgt)
            # img.anchor(ws.cell(row=25, column=22))
            ws.add_image(img, 'G25')
            print('[+] Image {0} written into {1}'.format(imgp + ost + imgt, xlsp + ost + xlst))
            logging.info('[+] Изображение {0} записано в {1}'.format(imgp + ost + imgt, xlsp + ost + xlst))
            wb.save(xlsp + ost + xlst)
            wb.close()
        except Exception as err:
            logging.error('[-] Произошла ошибка в модуле write_image: {0}'.format(str(err)))
            exit(0)


def write_sql(config: dict):
    """Записать результаты SQL запросов из DBData.json в Excel таблицу"""
    try:
        sqlp = config['PATH']['SQL_FILE'] + 'DBData.json'
        xlsp = config['PATH']['XLS_FILE']
        xlst = config['TYPE']['EXCEL']
        with open(sqlp) as sql_file:
            sql = json.load(sql_file)
        for ost in config['OST']:
                wb = load_workbook(xlsp + ost + xlst)
                w = wb.active
                w['G3'] = sql[ost]['ABBYYDB_BATCHPARAM']
                w['G22'] = sql[ost]['OSDB_QUEUEITEM']
                w['G24'] = sql[ost]['OSDB_CONDUCTOR']
                wb.save(xlsp + ost + xlst)
                wb.close()
                print('[+] SQL-queues {0} written into {1}'.format(ost, xlsp + ost + xlst))
                logging.info('[+] Результаты SQL-запросов {0} записаны в {1}'.format(ost, xlsp + ost + xlst))
    except Exception as err:
        logging.error('[-] Произошла ошибка в модуле write_sql: {0}'.format(str(err)))
        exit(0)


def main():
    print('[**] Data in excel writing started')
    logging.info('[**] Запись данных в excel-таблицу начата')
    config = read_config('config.json')
    # write_image(config)
    write_win_size(config)
    write_sql(config)
    print('[**] Data in excel writing finished')
    logging.info('[**] Запись данных в excel-таблицу окончена')


if __name__ == '__main__':
    main()
