import json
from openpyxl import load_workbook, drawing
import openpyxl
from openpyxl.drawing.image import Image
import logging

logging.basicConfig(filename='auto_check.log', format='%(asctime)s - %(levelname)s : %(message)s', level=logging.INFO)


def read_config(path_to_config: str) -> dict:
    """Считать данные из config.json"""
    try:
        with open(path_to_config) as data_file:
            data = json.load(data_file)
        print('[+] Reading config file')
        logging.info('[+] Конфигурационный фаил прочитан')
        return data
    except Exception as err:
        logging.error('[-] Произошла ошибка в модуле read_config: {0}'.format(str(err)))
        exit(0)


def write_image(config: dict):
    """Записать PNG фаил из папки со скриншотами в excel таблицу"""
    for ost in config['OST']:
        try:
            xlsp = config['PATH']['XLS_FILE']
            xlst = config['TYPE']['EXCEL']
            imgp = config['PATH']['VWTOOL_SCREENSHOT']
            imgt = config['TYPE']['INPUT_IMAGE']
            wb = load_workbook(xlsp + ost + xlst)
            ws = wb.active
            img = openpyxl.drawing.image.Image(imgp + ost + imgt)
            #img.anchor(ws.cell(row=25, column=22))
            ws.add_image(img, 'G25')
            print('[+] Image {0} written into {1}'.format(imgp + ost + imgt, xlsp + ost + xlst))
            logging.info('[+] Изображение {0} записано в {1}'.format(imgp + ost + imgt, xlsp + ost + xlst))
            wb.save(xlsp + ost + xlst)
            wb.close()
        except Exception as err:
            logging.error('[-] Произошла ошибка в модуле write_image: {0}'.format(str(err)))
            exit(0)


def main():
    print('[**] Start image-kostyling')
    logging.info('[**] Запись изображения при помощи костыля начата')
    config = read_config('config.json')
    write_image(config)
    print('[**] Stop image-kostyling')
    logging.info('[**] Запись изображения при помощи костыля закончена')


if __name__ == '__main__':
    main()
